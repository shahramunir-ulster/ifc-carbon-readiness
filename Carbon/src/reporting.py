from __future__ import annotations

from io import BytesIO

import pandas as pd


def filter_element_dataframe(
    df: pd.DataFrame,
    ifc_class: str | None = None,
    material: str | None = None,
    storey: str | None = None,
    readiness: str | None = None,
    search: str | None = None,
) -> pd.DataFrame:
    """Filter element result data for the dissertation-style dashboard."""
    if df is None or df.empty:
        return pd.DataFrame(columns=df.columns if hasattr(df, 'columns') else [])

    filtered = df.copy()

    if ifc_class:
        filtered = filtered[filtered['ifc_class'].astype(str).str.strip() == str(ifc_class).strip()]
    if material:
        filtered = filtered[filtered['material_name'].fillna('').astype(str).str.lower().str.contains(str(material).strip().lower(), na=False)]
    if storey:
        filtered = filtered[filtered['storey'].fillna('').astype(str).str.lower().str.contains(str(storey).strip().lower(), na=False)]
    if readiness:
        filtered = filtered[filtered['readiness_status'].fillna('').astype(str).str.strip() == str(readiness).strip()]
    if search:
        term = str(search).strip().lower()
        if term:
            searchable = filtered.astype(str).apply(lambda row: ' '.join(row.values).lower(), axis=1)
            filtered = filtered[searchable.str.contains(term, na=False)]

    return filtered.reset_index(drop=True)


def build_excel_export(
    result_df: pd.DataFrame,
    missing_df: pd.DataFrame,
    match_df: pd.DataFrame,
    summary_rows: list[dict],
    scope_df: pd.DataFrame | None = None,
) -> bytes:
    """Create a multi-sheet workbook export for dissertation reporting."""
    from openpyxl import Workbook
    from openpyxl.styles import Font

    workbook = Workbook()
    workbook.remove(workbook.active)

    def normalise_cell_value(value):
        if isinstance(value, list):
            return '; '.join(str(item) for item in value)
        if isinstance(value, tuple):
            return '; '.join(str(item) for item in value)
        return value

    def add_sheet(name: str, data: pd.DataFrame, freeze_top: bool = True):
        ws = workbook.create_sheet(title=name)
        if data.empty:
            ws.append(['No data available'])
            ws['A1'].font = Font(bold=True)
            return

        columns = list(data.columns)
        ws.append(columns)
        for row in data.itertuples(index=False, name=None):
            normalised_row = [normalise_cell_value(item) for item in row]
            ws.append(normalised_row)

        if freeze_top:
            ws.freeze_panes = 'A2'

        for cell in ws[1]:
            cell.font = Font(bold=True)

    add_sheet('Results', result_df)
    add_sheet('Missing Data', missing_df)
    add_sheet('Factor Matching', match_df)
    add_sheet('Summary', pd.DataFrame(summary_rows))
    if scope_df is not None:
        add_sheet('IFC Scope', scope_df)

    stream = BytesIO()
    workbook.save(stream)
    return stream.getvalue()
