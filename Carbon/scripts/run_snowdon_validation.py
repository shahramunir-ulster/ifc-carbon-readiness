from __future__ import annotations

from collections import Counter
from datetime import datetime, timezone
import hashlib
import math
from pathlib import Path
import sys

import pandas as pd
import streamlit
import ifcopenshell
import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill


ROOT = Path(__file__).resolve().parents[1]
if str(ROOT) not in sys.path:
    sys.path.insert(0, str(ROOT))

from src.assessment import process_element_record
from src.carbon_factors import load_carbon_factor_csv
from src.data_readiness import build_processing_error_record, compute_readiness_metrics
from src.ifc_parser import parse_ifc_file, validate_ifc_file
from src.version import APPLICATION_VERSION, source_tree_sha256


MODELS = {
    'Architectural': ROOT / 'Snowdon Towers Sample Architectural(1).ifc',
    'Structural': ROOT / 'Snowdon Towers Sample Structural(1).ifc',
}
FACTOR_FILE = ROOT / 'snowdon_controlled_factors.csv'
OUTPUT_FILE = ROOT / 'Snowdon_Formal_Validation.xlsx'
KNOWN_UNIT_FACTORS = {
    'm': 1.0,
    'm2': 1.0,
    'm3': 1.0,
    'kg': 1.0,
    'item': 1.0,
    'ft': 0.3048,
    'ft2': 0.09290304,
    'ft3': 0.028316846592,
    'in': 0.0254,
    'in2': 0.00064516,
    'in3': 0.000016387064,
}


def sha256(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open('rb') as stream:
        for chunk in iter(lambda: stream.read(1024 * 1024), b''):
            digest.update(chunk)
    return digest.hexdigest()


def flatten(value):
    if isinstance(value, list):
        return '; '.join(str(item) for item in value)
    if isinstance(value, dict):
        return '; '.join(f'{key}={item}' for key, item in value.items())
    return value


def result_frame(rows: list[dict]) -> pd.DataFrame:
    columns = [
        'model', 'global_id', 'ifc_class', 'element_type', 'element_name', 'object_type', 'storey',
        'material_name', 'material_normalised', 'multiple_materials',
        'material_match_status', 'matched_material_name', 'matching_method',
        'matching_confidence', 'material_id', 'factor_category', 'factor_notes',
        'factor_is_proxy', 'quantity_type', 'quantity_set_name', 'quantity_name', 'raw_quantity_value',
        'raw_quantity_unit', 'unit_conversion_factor', 'quantity_value',
        'quantity_unit', 'unit_source', 'quantity_supported', 'unit_compatible',
        'quantity_selection_method', 'density_kg_per_m3',
        'density_source', 'calculated_mass_kg', 'carbon_factor',
        'carbon_factor_unit', 'life_cycle_stage', 'carbon_factor_source',
        'calculation_possible', 'embodied_carbon_kgco2e', 'readiness_status',
        'issues', 'calculation_reason', 'material_extraction_error',
        'quantity_extraction_error', 'extraction_error', 'processing_error',
        'calculation_failure_code', 'carbon_factor_source_type', 'recommended_action',
    ]
    return pd.DataFrame([{key: flatten(row.get(key, '')) for key in columns} for row in rows])


def independently_validate(row: dict) -> dict:
    raw_value = float(row['raw_quantity_value'])
    raw_unit = str(row['raw_quantity_unit'])
    if raw_unit not in KNOWN_UNIT_FACTORS:
        raise ValueError(f'No independent conversion constant for {raw_unit!r}')
    independent_si = raw_value * KNOWN_UNIT_FACTORS[raw_unit]
    quantity_unit = str(row['quantity_unit'])
    factor_unit = str(row['carbon_factor_unit']).lower()
    density = float(row['density_kg_per_m3']) if pd.notna(row['density_kg_per_m3']) else None
    carbon_factor = float(row['carbon_factor'])

    if quantity_unit == 'm3' and factor_unit == 'kgco2e/kg':
        independent_mass = independent_si * density
        independent_carbon = independent_mass * carbon_factor
        formula = 'raw volume x unit conversion x density x carbon factor'
    elif quantity_unit == 'kg' and factor_unit == 'kgco2e/kg':
        independent_mass = independent_si
        independent_carbon = independent_mass * carbon_factor
        formula = 'raw mass x unit conversion x carbon factor'
    elif quantity_unit == 'm3' and factor_unit == 'kgco2e/m3':
        independent_mass = None
        independent_carbon = independent_si * carbon_factor
        formula = 'raw volume x unit conversion x carbon factor'
    elif quantity_unit == 'm2' and factor_unit == 'kgco2e/m2':
        independent_mass = None
        independent_carbon = independent_si * carbon_factor
        formula = 'raw area x unit conversion x carbon factor'
    elif quantity_unit == 'item' and factor_unit == 'kgco2e/item':
        independent_mass = None
        independent_carbon = independent_si * carbon_factor
        formula = 'item count x carbon factor'
    else:
        raise ValueError(f'Unsupported validation combination: {quantity_unit}, {factor_unit}')

    prototype_si = float(row['quantity_value'])
    prototype_carbon = float(row['embodied_carbon_kgco2e'])
    abs_difference = abs(prototype_carbon - independent_carbon)
    percent_difference = abs_difference / independent_carbon * 100 if independent_carbon else 0.0
    return {
        **{key: row.get(key) for key in (
            'model', 'global_id', 'ifc_class', 'element_name', 'storey',
            'material_name', 'matched_material_name', 'quantity_name',
            'raw_quantity_value', 'raw_quantity_unit', 'quantity_value',
            'quantity_unit', 'density_kg_per_m3', 'carbon_factor',
            'carbon_factor_unit', 'life_cycle_stage',
        )},
        'independent_conversion_factor': KNOWN_UNIT_FACTORS[raw_unit],
        'independent_si_quantity': independent_si,
        'si_quantity_difference': prototype_si - independent_si,
        'independent_mass_kg': independent_mass,
        'prototype_mass_kg': row.get('calculated_mass_kg'),
        'independent_carbon_kgco2e': independent_carbon,
        'prototype_carbon_kgco2e': prototype_carbon,
        'absolute_difference_kgco2e': abs_difference,
        'percent_difference': percent_difference,
        'validation_result': 'PASS' if math.isclose(prototype_carbon, independent_carbon, rel_tol=1e-9, abs_tol=1e-6) else 'FAIL',
        'independent_formula': formula,
    }


def choose_sample(calculated: list[dict], limit: int = 20) -> list[dict]:
    ordered = sorted(
        calculated,
        key=lambda row: (
            row['model'], row.get('ifc_class', ''), row.get('material_normalised', ''),
            -float(row.get('embodied_carbon_kgco2e') or 0), row.get('global_id', ''),
        ),
    )
    selected = []
    selected_ids = set()
    group_counts: Counter[tuple] = Counter()
    known_id = '1Z_D8qyun52938MRTkC1du'
    known = next((row for row in calculated if row.get('global_id') == known_id), None)
    if known:
        selected.append(known)
        selected_ids.add(known.get('global_id'))
        group_counts[(known['model'], known.get('ifc_class'), known.get('material_normalised'))] += 1
    contributing_materials = sorted({
        str(row.get('matched_material_name') or '').strip()
        for row in calculated if float(row.get('embodied_carbon_kgco2e') or 0) > 0
    } - {''})
    for material in contributing_materials:
        if len(selected) >= limit:
            break
        if any(str(row.get('matched_material_name') or '').strip() == material for row in selected):
            continue
        candidates = [
            row for row in calculated
            if str(row.get('matched_material_name') or '').strip() == material
            and row.get('global_id') not in selected_ids
        ]
        if candidates:
            row = max(candidates, key=lambda item: float(item.get('embodied_carbon_kgco2e') or 0))
            selected.append(row)
            selected_ids.add(row.get('global_id'))
            group_counts[(row['model'], row.get('ifc_class'), row.get('material_normalised'))] += 1
    for row in ordered:
        if len(selected) >= limit:
            break
        if known and row.get('global_id') == known_id:
            continue
        group = (row['model'], row.get('ifc_class'), row.get('material_normalised'))
        if group_counts[group] >= 1:
            continue
        selected.append(row)
        selected_ids.add(row.get('global_id'))
        group_counts[group] += 1
    if len(selected) < limit:
        for row in ordered:
            if len(selected) >= limit:
                break
            if row.get('global_id') not in selected_ids:
                selected.append(row)
                selected_ids.add(row.get('global_id'))
    return selected


def format_workbook(path: Path) -> None:
    from openpyxl import load_workbook

    workbook = load_workbook(path)
    header_fill = PatternFill('solid', fgColor='1F4E78')
    for worksheet in workbook.worksheets:
        worksheet.freeze_panes = 'A2'
        worksheet.auto_filter.ref = worksheet.dimensions
        for cell in worksheet[1]:
            cell.fill = header_fill
            cell.font = Font(color='FFFFFF', bold=True)
            cell.alignment = Alignment(wrap_text=True, vertical='top')
        for column_cells in worksheet.columns:
            values = [str(cell.value or '') for cell in list(column_cells)[:250]]
            width = min(max(max((len(value) for value in values), default=8) + 2, 10), 45)
            worksheet.column_dimensions[column_cells[0].column_letter].width = width
        for row in worksheet.iter_rows(min_row=2):
            for cell in row:
                cell.alignment = Alignment(vertical='top', wrap_text=False)
    workbook.save(path)


def main() -> None:
    factors = load_carbon_factor_csv(FACTOR_FILE)
    all_results: list[dict] = []
    all_scope: list[dict] = []
    model_summaries = []

    for model_name, path in MODELS.items():
        validation = validate_ifc_file(path)
        if not validation['valid']:
            raise ValueError(f'{path.name}: {validation["reason"]}')
        parsed = parse_ifc_file(path)
        processed = []
        processing_error_count = 0
        for element in parsed['elements']:
            try:
                result = process_element_record(element.copy(), factors)
            except Exception as exc:
                result = build_processing_error_record(element, exc)
                processing_error_count += 1
            result['model'] = model_name
            processed.append(result)
        all_results.extend(processed)
        all_scope.extend([{'model': model_name, **row} for row in parsed['scope_report']])
        metrics = compute_readiness_metrics(processed)
        partial_total = sum(row['embodied_carbon_kgco2e'] for row in processed if row['calculation_possible'])
        model_summaries.append({
            'model': model_name,
            'file': path.name,
            'sha256': sha256(path),
            'IFC_validation': validation['reason'],
            'project_name': parsed['metadata']['project_name'],
            'elements_scanned': parsed['total_ifc_elements'],
            'assessment_eligible': parsed['in_scope_elements'],
            'review_required': parsed['review_required_elements'],
            'not_applicable': parsed['not_applicable_elements'],
            'extraction_error_count': parsed['error_count'],
            'processing_error_count': processing_error_count,
            **metrics,
            'partial_A1_A3_kgCO2e': partial_total,
            'partial_A1_A3_tCO2e': partial_total / 1000,
        })

    results_df = result_frame(all_results)
    calculated = [row for row in all_results if row.get('calculation_possible')]
    validation_rows = [independently_validate(row) for row in choose_sample(calculated)]
    validation_df = pd.DataFrame(validation_rows)

    material_coverage = (
        results_df.groupby(['model', 'material_name', 'material_normalised', 'material_match_status'], dropna=False)
        .agg(
            elements=('global_id', 'count'),
            calculated_elements=('calculation_possible', 'sum'),
            partial_A1_A3_kgCO2e=('embodied_carbon_kgco2e', 'sum'),
        )
        .reset_index()
        .sort_values(['model', 'elements'], ascending=[True, False])
    )
    readiness = (
        results_df.groupby(['model', 'readiness_status'], dropna=False)
        .size().rename('elements').reset_index()
        .sort_values(['model', 'elements'], ascending=[True, False])
    )
    methodology = pd.DataFrame([
        {'topic': 'Assessment boundary', 'statement': 'Product stage A1-A3 only.'},
        {'topic': 'Scope denominator', 'statement': 'Only assessment-eligible physical elements; ambiguous containers/composites are reported separately.'},
        {'topic': 'Quantity selection', 'statement': 'Net quantity preferred; gross quantity used only as fallback. A factor-compatible quantity is selected.'},
        {'topic': 'Imperial conversion', 'statement': 'ft to m = 0.3048; ft2 to m2 = 0.09290304; ft3 to m3 = 0.028316846592.'},
        {'topic': 'Volume-to-mass formula', 'statement': 'Mass (kg) = IFC volume converted to m3 x density (kg/m3).'},
        {'topic': 'Carbon formula', 'statement': 'A1-A3 kgCO2e = mass (kg) x carbon factor (kgCO2e/kg).'},
        {'topic': 'Material matching', 'statement': 'Controlled exact match after conservative normalisation; multi-material objects are not automatically allocated.'},
        {'topic': 'Factor coverage', 'statement': 'Matched elements divided by every assessment-eligible element; this measures overall factor coverage, not matching-algorithm success.'},
        {'topic': 'Factor match success', 'statement': 'Matched divided by matched plus genuinely unmatched single-material elements; missing, invalid, and multi-material cases are outside this denominator.'},
        {'topic': 'Reported total', 'statement': 'Partial prototype estimate only: unmatched, multi-material, missing-quantity, and excluded elements are not included.'},
        {'topic': 'Independent check', 'statement': '20 representative calculated elements are recomputed using hard-coded published unit constants and the controlled factor values.'},
        {'topic': 'Validation limitation', 'statement': 'Arithmetic agreement does not validate IFC geometry, model authoring, material specifications, density proxies, or ICE factor applicability.'},
        {'topic': 'Required external check', 'statement': 'Compare selected quantities with native Revit schedules and replace proxies with verified project EPD/product data before decision use.'},
    ])
    run_info = pd.DataFrame([
        {'item': 'Generated UTC', 'value': datetime.now(timezone.utc).isoformat()},
        {'item': 'Application version', 'value': APPLICATION_VERSION},
        {'item': 'Source code SHA-256', 'value': source_tree_sha256(ROOT)},
        {'item': 'Python version', 'value': sys.version.split()[0]},
        {'item': 'IfcOpenShell version', 'value': getattr(ifcopenshell, 'version', 'unknown')},
        {'item': 'pandas version', 'value': pd.__version__},
        {'item': 'Streamlit version', 'value': streamlit.__version__},
        {'item': 'openpyxl version', 'value': openpyxl.__version__},
        {'item': 'Factor file', 'value': FACTOR_FILE.name},
        {'item': 'Factor file SHA-256', 'value': sha256(FACTOR_FILE)},
        {'item': 'Independent sample size', 'value': len(validation_df)},
        {'item': 'Independent sample passes', 'value': int((validation_df['validation_result'] == 'PASS').sum())},
        {'item': 'Independent sample failures', 'value': int((validation_df['validation_result'] == 'FAIL').sum())},
    ])

    scope_df = pd.DataFrame(all_scope)
    factor_export = factors.drop(columns=['is_demo_value'], errors='ignore')
    with pd.ExcelWriter(OUTPUT_FILE, engine='openpyxl') as writer:
        run_info.to_excel(writer, sheet_name='Run Information', index=False)
        pd.DataFrame(model_summaries).to_excel(writer, sheet_name='Model Summary', index=False)
        methodology.to_excel(writer, sheet_name='Methodology', index=False)
        validation_df.to_excel(writer, sheet_name='Arithmetic Verification', index=False)
        factor_export.to_excel(writer, sheet_name='Controlled Factors', index=False)
        material_coverage.to_excel(writer, sheet_name='Material Coverage', index=False)
        readiness.to_excel(writer, sheet_name='Readiness Status', index=False)
        results_df.to_excel(writer, sheet_name='Eligible Element Results', index=False)
        scope_df.to_excel(writer, sheet_name='IFC Scope Register', index=False)
    format_workbook(OUTPUT_FILE)

    summary = pd.DataFrame(model_summaries)[[
        'model', 'elements_scanned', 'assessment_eligible', 'calculated_elements',
        'calculation_success_rate', 'quantity_completeness', 'factor_coverage_rate',
        'factor_match_success_rate',
        'missing_or_incomplete', 'unmatched_materials', 'multiple_material_elements',
        'incompatible_units', 'review_required', 'extraction_error_count',
        'processing_error_count',
        'partial_A1_A3_tCO2e',
    ]]
    print(summary.to_string(index=False))
    print(f'Independent arithmetic verification: {(validation_df["validation_result"] == "PASS").sum()}/{len(validation_df)} passed')
    print(f'Report: {OUTPUT_FILE}')


if __name__ == '__main__':
    main()
