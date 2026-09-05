from __future__ import annotations

import pandas as pd
from io import BytesIO
from pathlib import Path
from openpyxl import load_workbook

from src.reporting import build_excel_export, filter_element_dataframe


def test_filter_element_dataframe_supports_status_and_search():
    df = pd.DataFrame([
        {
            'global_id': 'A1',
            'ifc_class': 'IfcWall',
            'element_name': 'Wall 01',
            'storey': 'Level 1',
            'material_name': 'Concrete',
            'readiness_status': 'ready / carbon calculated',
            'issues': [],
        },
        {
            'global_id': 'A2',
            'ifc_class': 'IfcBeam',
            'element_name': 'Beam 01',
            'storey': 'Level 2',
            'material_name': 'Steel',
            'readiness_status': 'missing material',
            'issues': ['missing material'],
        },
    ])

    filtered = filter_element_dataframe(
        df,
        ifc_class='IfcWall',
        material='Concrete',
        storey='Level 1',
        readiness='ready / carbon calculated',
        search='Wall',
    )

    assert len(filtered) == 1
    assert filtered.iloc[0]['global_id'] == 'A1'


def test_build_excel_export_creates_workbook_bytes():
    df = pd.DataFrame([
        {
            'global_id': 'A1',
            'ifc_class': 'IfcWall',
            'material_name': 'Concrete',
            'embodied_carbon_kgco2e': 12.5,
            'readiness_status': 'ready / carbon calculated',
        }
    ])
    missing_df = pd.DataFrame([
        {'global_id': 'A2', 'ifc_class': 'IfcSlab', 'issues': ['missing material']}
    ])
    match_df = pd.DataFrame([
        {'global_id': 'A1', 'material_name': 'Concrete', 'material_match_status': 'matched'}
    ])
    summary_rows = [{'metric': 'total_elements', 'value': 1}]

    payload = build_excel_export(df, missing_df, match_df, summary_rows)

    assert isinstance(payload, bytes)
    assert payload.startswith(b'PK')


def test_excel_export_includes_ifc_scope_audit_sheet():
    scope_df = pd.DataFrame([{
        'global_id': 'P1', 'ifc_class': 'IfcBuildingElementProxy',
        'scope_status': 'review_required', 'scope_reason': 'Ambiguous class',
    }])
    payload = build_excel_export(pd.DataFrame(), pd.DataFrame(), pd.DataFrame(), [], scope_df)
    workbook = load_workbook(BytesIO(payload), read_only=True)
    assert 'IFC Scope' in workbook.sheetnames


def test_app_excel_export_is_wired_to_complete_not_filtered_frames():
    source = (Path(__file__).resolve().parents[1] / 'app.py').read_text(encoding='utf-8')
    start = source.index('excel_bytes = build_excel_export(')
    end = source.index("st.download_button(\n        'Download Excel workbook'", start)
    call = source[start:end]
    assert 'result_table.reset_index' in call
    assert 'missing_report.reset_index' in call
    assert 'match_df.reset_index' in call
    assert 'filtered_results' not in call
    assert 'filtered_missing' not in call
    assert 'filtered_match' not in call


def test_app_persists_assessment_result_for_filter_reruns():
    source = (Path(__file__).resolve().parents[1] / 'app.py').read_text(encoding='utf-8')
    assert 'st.session_state.assessment_result = result' in source
    assert "assessment_result = st.session_state.assessment_result" in source
    assert 'def run_live_assessment(ifc_bytes: bytes | memoryview, factor_bytes: bytes)' in source
    assert '@st.cache_data(show_spinner=False, max_entries=2, ttl=3600)' not in source
    assert 'uploaded_file.getbuffer()' in source
    assert "input_signature = 'mode:demo' if is_demo_run else 'mode:live'" in source
    assert 'st.session_state.assessment_excel_bytes' in source
